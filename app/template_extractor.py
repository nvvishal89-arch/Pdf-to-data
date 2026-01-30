"""
Template extractor: reads Sample format SQ.xlsx and outputs reusable template config
(anchors + column mapping) for the PDF pipeline.
"""
import json
from pathlib import Path

from openpyxl import load_workbook

from app.template_config import (
    TemplateConfig,
    HeaderAnchor,
    TableColumn,
)

# Expected header labels (spec) -> schema key
HEADER_LABEL_TO_KEY = {
    "company name": "company_name",
    "project name": "project_name",
    "client name": "client_name",
    "quotation no": "quotation_no",
    "quotation no.": "quotation_no",
    "date": "date",
    "prepared by": "prepared_by",
}

# Table column header (normalized) -> schema key
TABLE_HEADER_TO_KEY = {
    "s.no": "sr_no",
    "s.no.": "sr_no",
    "sr no": "sr_no",
    "product description": "name",
    "description": "name",
    "size": "dimensions",
    "dimensions": "dimensions",
    "size / dimensions": "dimensions",
    "area": "area",
    "material": "material",
    "finish": "finish",
    "qty": "qty",
    "quantity": "qty",
    "rate": "unit_price",
    "amount": "amount",
    "reference image": "images",
    "reference image(s)": "images",
    "image": "images",
}


def _normalize(s: str | None) -> str:
    if s is None:
        return ""
    return str(s).strip().lower()


def _find_header_block(ws, max_rows: int = 20) -> list[tuple[int, int, str, str]]:
    """Scan top rows for header labels; return (row, col, label, key)."""
    found: list[tuple[int, int, str, str]] = []
    for row_idx in range(1, max_rows + 1):
        for col_idx in range(1, 25):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = _normalize(cell.value)
            if not val:
                continue
            for label, key in HEADER_LABEL_TO_KEY.items():
                if label in val or val in label:
                    found.append((row_idx, col_idx, val, key))
                    break
    return found


def _find_table_header_row(ws, max_rows: int = 30) -> int:
    """Return 1-based row index of the row containing 'S.No' or similar."""
    for row_idx in range(1, max_rows + 1):
        for col_idx in range(1, 20):
            val = _normalize(ws.cell(row=row_idx, column=col_idx).value)
            if val in ("s.no", "s.no.", "sr no", "sr.no"):
                return row_idx
    return 0


def _find_table_columns(ws, header_row: int) -> list[tuple[int, str, str]]:
    """Return (col_index, header_text, key) for each table column."""
    result: list[tuple[int, str, str]] = []
    for col_idx in range(1, 20):
        val = _normalize(ws.cell(row=header_row, column=col_idx).value)
        if not val:
            continue
        for hdr, key in TABLE_HEADER_TO_KEY.items():
            if hdr in val or val in hdr:
                result.append((col_idx, val, key))
                break
    return result


def extract_template(excel_path: str | Path) -> TemplateConfig:
    """
    Parse Sample format SQ.xlsx and build TemplateConfig.
    """
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Template Excel not found: {path}")

    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        raise ValueError("Workbook has no active sheet")

    # Header block
    header_cells = _find_header_block(ws)
    anchors: list[HeaderAnchor] = []
    for row, col, label, key in header_cells:
        value_col = col + 1
        anchors.append(
            HeaderAnchor(label=label, key=key, row=row, col=col, value_col=value_col)
        )

    # Table
    table_header_row = _find_table_header_row(ws)
    table_columns: list[TableColumn] = []
    if table_header_row:
        for col_index, header_text, key in _find_table_columns(ws, table_header_row):
            table_columns.append(
                TableColumn(header=header_text, key=key, col_index=col_index)
            )
    data_start_row = table_header_row + 1 if table_header_row else 1

    wb.close()

    return TemplateConfig(
        header_anchors=anchors,
        table_header_row=table_header_row,
        table_columns=table_columns,
        data_start_row=data_start_row,
    )


def extract_template_to_json(excel_path: str | Path, output_path: str | Path | None = None) -> str:
    """
    Extract template config and return JSON string; optionally write to file.
    """
    config = extract_template(excel_path)
    json_str = config.model_dump_json(indent=2)
    if output_path:
        Path(output_path).write_text(json_str, encoding="utf-8")
    return json_str


def load_template_config(config_path: str | Path) -> TemplateConfig:
    """Load TemplateConfig from a JSON file."""
    path = Path(config_path)
    data = json.loads(path.read_text(encoding="utf-8"))
    return TemplateConfig.model_validate(data)
